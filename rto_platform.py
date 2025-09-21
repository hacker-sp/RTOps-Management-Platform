#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Red Team Operations Platform — single-file Flask + SQLite
# Now with:
# - Full ATT&CK import + robust XLSX enrichment (names/descriptions per tactic)
# - Better MITRE Map (tactics, search)
# - People & Assignments (key stakeholders + duties per Exercise)
# - Dashboards (Chart.js), Kill Chain, Reports (PDF), RoE, Emulation Plans, Findings, IoCs, Intel

import os, sqlite3, json, datetime
from collections import defaultdict
from io import BytesIO
from flask import Flask, g, request, redirect, url_for, send_file, make_response

APP_NAME = "RTOps"
DB_PATH = os.path.join(os.path.dirname(__file__), "rtops.sqlite3")

MITRE_JSON_CANDIDATES = [
    os.path.join(os.getcwd(), "enterprise-attack.json"),
    "/mnt/data/enterprise-attack.json",
]
MITRE_XLSX_CANDIDATES = [
    os.path.join(os.getcwd(), "enterprise-attack-v17.1.xlsx"),
    "/mnt/data/enterprise-attack-v17.1.xlsx",
]

# Canonical ATT&CK Enterprise tactic order
TACTIC_ORDER = [
    "reconnaissance",
    "resource-development",
    "initial-access",
    "execution",
    "persistence",
    "privilege-escalation",
    "defense-evasion",
    "credential-access",
    "discovery",
    "lateral-movement",
    "collection",
    "command-and-control",
    "exfiltration",
    "impact",
]
TACTIC_TITLES = {
    "reconnaissance": "Reconnaissance",
    "resource-development": "Resource Development",
    "initial-access": "Initial Access",
    "execution": "Execution",
    "persistence": "Persistence",
    "privilege-escalation": "Privilege Escalation",
    "defense-evasion": "Defense Evasion",
    "credential-access": "Credential Access",
    "discovery": "Discovery",
    "lateral-movement": "Lateral Movement",
    "collection": "Collection",
    "command-and-control": "Command & Control",
    "exfiltration": "Exfiltration",
    "impact": "Impact",
}

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.units import cm

app = Flask(__name__)

# --------------------- DB helpers ---------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db:
        db.close()

def init_db():
    db = get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS roe (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT, customer TEXT, rules TEXT, start_date TEXT, end_date TEXT,
        contacts TEXT, approvals TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS emulation_plans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT, threat_actor TEXT, scope TEXT, objectives TEXT,
        ttps_json TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS exercises (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT, description TEXT, env TEXT, start_date TEXT, end_date TEXT,
        objectives TEXT, status TEXT, related_roe_id INTEGER, related_plan_id INTEGER,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS findings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        exercise_id INTEGER, title TEXT, severity TEXT, asset TEXT,
        description TEXT, evidence TEXT, remediation TEXT,
        owner TEXT, status TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS iocs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT, value TEXT, related_technique TEXT, notes TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS intel (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT, source TEXT, tags TEXT, notes TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS ttps (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        technique_id TEXT, tactic TEXT, name TEXT, description TEXT, refs TEXT, created_at TEXT,
        UNIQUE(technique_id, tactic) ON CONFLICT IGNORE
    );
    CREATE TABLE IF NOT EXISTS killchain (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT, mapping_json TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY, value TEXT
    );
    -- People & Assignments
    CREATE TABLE IF NOT EXISTS people (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT, role TEXT, contact TEXT, notes TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS assignments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        exercise_id INTEGER, person_id INTEGER, duty TEXT, created_at TEXT
    );
    """)
    db.commit()

def get_setting(key, default=None):
    db = get_db()
    row = db.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return row["value"] if row else default

def set_setting(key, value):
    db = get_db()
    db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, value))
    db.commit()

@app.before_request
def ensure_db():
    if not os.path.exists(DB_PATH):
        open(DB_PATH, "a").close()
    init_db()

# --------------------- Seed fallback TTPs ---------------------
DEFAULT_TTPS = [
    {"technique_id": "T1059", "tactic": "execution", "name": "Command and Scripting Interpreter",
     "description": "Execute commands and scripts via shells/interpreters.", "refs": "https://attack.mitre.org/techniques/T1059/"},
    {"technique_id": "T1021", "tactic": "lateral-movement", "name": "Remote Services",
     "description": "RDP/SMB/SSH for lateral movement.", "refs": "https://attack.mitre.org/techniques/T1021/"},
    {"technique_id": "T1003", "tactic": "credential-access", "name": "OS Credential Dumping",
     "description": "Dump creds from OS components.", "refs": "https://attack.mitre.org/techniques/T1003/"},
]
def seed_minimal_ttps():
    db = get_db()
    c = db.execute("SELECT COUNT(*) c FROM ttps").fetchone()["c"]
    if c == 0:
        now = datetime.datetime.utcnow().isoformat()
        for t in DEFAULT_TTPS:
            db.execute("""INSERT INTO ttps (technique_id,tactic,name,description,refs,created_at)
                          VALUES (?,?,?,?,?,?)""",
                       (t["technique_id"], t["tactic"], t["name"], t["description"], t["refs"], now))
        db.commit()

@app.before_request
def maybe_seed():
    seed_minimal_ttps()

# --------------------- HTML scaffolding ---------------------
def html_page(title, body_html, extra_head=""):
    return f"""<!DOCTYPE html>
<html lang="en" class="h-full" data-theme="light">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>{APP_NAME} • {title}</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <script src="https://unpkg.com/htmx.org@1.9.12"></script>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
  <script>
    function applyTheme(t) {{
      document.documentElement.setAttribute('data-theme', t);
      document.documentElement.classList.toggle('dark', t==='dark');
      localStorage.setItem('rtops-theme', t);
    }}
    document.addEventListener('DOMContentLoaded', function() {{
      const saved = localStorage.getItem('rtops-theme') || 'dark';
      applyTheme(saved);
    }});
    function toggleTheme() {{
      const cur = localStorage.getItem('rtops-theme') || 'dark';
      applyTheme(cur === 'dark' ? 'light' : 'dark');
    }}
  </script>
  <style>
    :root {{ color-scheme: light dark; }}
    .chip {{ padding:.15rem .5rem; border-radius:.5rem; font-size:.75rem; }}
    .card {{ border:1px solid rgba(100,100,100,.2); border-radius:1rem; padding:1rem; }}
    .grid-2 {{ display:grid; grid-template-columns: 1fr 1fr; gap:1rem; }}
    .grid-3 {{ display:grid; grid-template-columns: repeat(3,1fr); gap:1rem; }}
    .grid-4 {{ display:grid; grid-template-columns: repeat(4,1fr); gap:1rem; }}
    .dark .card {{ border-color: rgba(255,255,255,.15); }}
    .btn {{ padding:.55rem .9rem; border-radius:.75rem; border:1px solid rgba(120,120,120,.25); }}
    .btn:hover {{ filter:brightness(0.98); }}
    .tbl th, .tbl td {{ padding:.5rem; border-bottom:1px solid rgba(127,127,127,.2); text-align:left; }}
    .search {{ border:1px solid rgba(120,120,120,.25); padding:.5rem .75rem; border-radius:.5rem; width:100%; }}
    .pill {{ border:1px solid rgba(127,127,127,.25); padding:.25rem .5rem; border-radius:.5rem; }}
  </style>
  {extra_head}
</head>
<body class="min-h-full bg-white text-slate-800 dark:bg-slate-900 dark:text-slate-100">
  <header class="sticky top-0 z-10 backdrop-blur bg-white/70 dark:bg-slate-900/70 border-b border-slate-200 dark:border-slate-800">
    <div class="max-w-7xl mx-auto p-3 flex items-center gap-3">
      <a href="/" class="font-bold text-lg">{APP_NAME}</a>
      <nav class="flex flex-wrap gap-2 text-sm">
        <a class="btn" href="/roe">RoE</a>
        <a class="btn" href="/emulations">Emulation Plans</a>
        <a class="btn" href="/exercises">Exercises</a>
        <a class="btn" href="/findings">Findings</a>
        <a class="btn" href="/iocs">IoCs</a>
        <a class="btn" href="/intel">Threat Intel</a>
        <a class="btn" href="/mitre">MITRE Map</a>
        <a class="btn" href="/killchain">Kill Chain</a>
        <a class="btn" href="/reports">Reports</a>
        <a class="btn" href="/people">People</a>
        <a class="btn" href="/admin">Admin</a>
      </nav>
      <div class="flex-1"></div>
      <button class="btn" onclick="toggleTheme()">Light/Dark</button>
    </div>
  </header>
  <main class="max-w-7xl mx-auto p-4">
    <h1 class="text-2xl font-semibold mb-4">{title}</h1>
    {body_html}
    <footer class="opacity-70 text-xs mt-8">Created by Joas A Santos • SQLite • {APP_NAME}</footer>
  </main>
</body>
</html>"""

def now_iso(): return datetime.datetime.utcnow().isoformat()
def rows_to_list(rows): return [dict(r) for r in rows]
def esc_attr(s: str) -> str:
    if s is None: return ""
    return str(s).replace("'", "&#39;").replace('"', "&quot;")

def kpi_cards():
    db = get_db()
    count = lambda q: db.execute(q).fetchone()["c"]
    c_ex  = count("SELECT COUNT(*) c FROM exercises")
    c_fn  = count("SELECT COUNT(*) c FROM findings")
    c_ioc = count("SELECT COUNT(*) c FROM iocs")
    c_plan= count("SELECT COUNT(*) c FROM emulation_plans")
    c_roe = count("SELECT COUNT(*) c FROM roe")
    c_people = count("SELECT COUNT(*) c FROM people")
    mitre_loaded = "Yes" if get_setting("mitre_loaded", "0") == "1" else "No"
    return f"""
    <div class="grid-4">
      <div class="card"><div class="text-sm">Exercises</div><div class="text-3xl font-bold">{c_ex}</div></div>
      <div class="card"><div class="text-sm">Findings</div><div class="text-3xl font-bold">{c_fn}</div></div>
      <div class="card"><div class="text-sm">IoCs</div><div class="text-3xl font-bold">{c_ioc}</div></div>
      <div class="card"><div class="text-sm">Emulation Plans</div><div class="text-3xl font-bold">{c_plan}</div></div>
      <div class="card"><div class="text-sm">RoE</div><div class="text-3xl font-bold">{c_roe}</div></div>
      <div class="card"><div class="text-sm">Stakeholders</div><div class="text-3xl font-bold">{c_people}</div></div>
      <div class="card"><div class="text-sm">ATT&CK Loaded</div><div class="text-3xl font-bold">{mitre_loaded}</div></div>
      <div class="card"><div class="text-sm">Reports</div><div class="text-lg">Exercise ▸ PDF</div></div>
    </div>
    """

def chart_block(canvas_id, title):
    return f"""
    <div class="card">
      <div class="text-sm opacity-80 mb-2">{title}</div>
      <canvas id="{canvas_id}" height="120"></canvas>
    </div>
    """

# --------------------- Home / Dashboard ---------------------
@app.get("/")
def home():
    db = get_db()
    kpis = kpi_cards()

    sev_rows = db.execute("SELECT severity, COUNT(*) c FROM findings GROUP BY severity").fetchall()
    sev_data = {r["severity"]: r["c"] for r in sev_rows}

    st_rows = db.execute("""
      SELECT substr(created_at,1,7) m, status, COUNT(*) c
      FROM findings GROUP BY m, status ORDER BY m ASC
    """).fetchall()
    months = sorted({r["m"] for r in st_rows})
    statuses = ["Open","In Progress","Closed"]
    series = {s: [0]*len(months) for s in statuses}
    m_index = {m:i for i,m in enumerate(months)}
    for r in st_rows:
        if r["status"] in series:
            series[r["status"]][m_index[r["m"]]] = r["c"]

    charts_js = f"""
    <script>
      const sevCtx = document.getElementById('sevChart');
      new Chart(sevCtx, {{
        type: 'bar',
        data: {{
          labels: {json.dumps(list(sev_data.keys()))},
          datasets: [{{ label: 'Findings', data: {json.dumps(list(sev_data.values()))} }}]
        }},
        options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }} }}
      }});

      const lineCtx = document.getElementById('statusChart');
      new Chart(lineCtx, {{
        type: 'line',
        data: {{
          labels: {json.dumps(months)},
          datasets: [
            {{ label: 'Open', data: {json.dumps(series['Open'])} }},
            {{ label: 'In Progress', data: {json.dumps(series['In Progress'])} }},
            {{ label: 'Closed', data: {json.dumps(series['Closed'])} }}
          ]
        }},
        options: {{ responsive: true }}
      }});
    </script>
    """

    body = f"""
    {kpis}
    <div class="grid-2 mt-6">
      {chart_block("sevChart", "Findings by Severity")}
      {chart_block("statusChart", "Findings Status by Month")}
    </div>
    """
    return html_page("Dashboard", body, extra_head=charts_js)

# --------- RoE ----------
@app.route("/roe", methods=["GET","POST"])
def roe():
    db = get_db()
    if request.method == "POST":
        f = request.form
        db.execute("""INSERT INTO roe (title,customer,rules,start_date,end_date,contacts,approvals,created_at)
                      VALUES (?,?,?,?,?,?,?,?)""",
                   (f.get("title"), f.get("customer"), f.get("rules"), f.get("start_date"), f.get("end_date"),
                    f.get("contacts"), f.get("approvals"), now_iso()))
        db.commit()
        return redirect(url_for("roe"))
    rows = rows_to_list(db.execute("SELECT * FROM roe ORDER BY id DESC"))
    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">Create RoE</div>
        <input name="title" placeholder="Title" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <input name="customer" placeholder="Customer / Stakeholder" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <div class="grid-2">
          <input name="start_date" type="date" class="p-2 rounded bg-transparent border"/>
          <input name="end_date" type="date" class="p-2 rounded bg-transparent border"/>
        </div>
        <textarea name="rules" placeholder="Rules of Engagement (scope, OOB, legal, comms)" class="w-full mb-2 p-2 rounded bg-transparent border h-28"></textarea>
        <textarea name="contacts" placeholder="POCs, escalation, channels" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <textarea name="approvals" placeholder="Approvals & signatures" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <button class="btn mt-2">Save</button>
      </form>
      <div class="card">
        <div class="text-lg font-semibold mb-2">RoE List</div>
        <table class="tbl w-full text-sm">
          <tr><th>ID</th><th>Title</th><th>Customer</th><th>Period</th></tr>
          {''.join(f"<tr><td>{r['id']}</td><td>{r['title']}</td><td>{r['customer']}</td><td>{r['start_date']} → {r['end_date']}</td></tr>" for r in rows)}
        </table>
      </div>
    </div>
    """
    return html_page("Rules of Engagement", body)

# --------- Emulation Plans ----------
@app.route("/emulations", methods=["GET","POST"])
def emulations():
    db = get_db()
    if request.method == "POST":
        f = request.form
        ttps_json = f.get("ttps_json") or "[]"
        db.execute("""INSERT INTO emulation_plans (title,threat_actor,scope,objectives,ttps_json,created_at)
                      VALUES (?,?,?,?,?,?)""",
                   (f.get("title"), f.get("threat_actor"), f.get("scope"), f.get("objectives"),
                    ttps_json, now_iso()))
        db.commit()
        return redirect(url_for("emulations"))
    plans = rows_to_list(db.execute("SELECT * FROM emulation_plans ORDER BY id DESC"))
    all_ttps = rows_to_list(db.execute("SELECT technique_id,tactic,name FROM ttps ORDER BY tactic,name"))
    ttp_opts = "".join([
        f"<option value='\"{t['technique_id']}|{t['tactic']}|{t['name'] or t['technique_id']}\"'>"
        f"{TACTIC_TITLES.get(t['tactic'], t['tactic'].title())} • {(t['name'] or t['technique_id'])} ({t['technique_id']})"
        f"</option>"
        for t in all_ttps
    ])
    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">New Emulation Plan</div>
        <input name="title" placeholder="Plan title" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <input name="threat_actor" placeholder="Threat actor / Campaign" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <textarea name="scope" placeholder="Scope" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <textarea name="objectives" placeholder="Objectives" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <div class="mb-2">
          <label class="text-sm">Select TTPs to include:</label>
          <select id="ttp_select" class="w-full p-2 rounded bg-transparent border" multiple size="8">
            {ttp_opts}
          </select>
          <input type="hidden" id="ttps_json" name="ttps_json"/>
          <button type="button" class="btn mt-2" onclick="
            const sel=[...document.getElementById('ttp_select').selectedOptions].map(o=>JSON.parse(o.value));
            document.getElementById('ttps_json').value = JSON.stringify(sel);
            alert('TTPs staged. Submit the form to save.');
          ">Stage TTPs</button>
        </div>
        <button class="btn mt-2">Save Plan</button>
      </form>
      <div class="card">
        <div class="text-lg font-semibold mb-2">Plans</div>
        <table class="tbl w-full text-sm">
          <tr><th>ID</th><th>Title</th><th>Actor</th><th>TTPs</th></tr>
          {''.join(f"<tr><td>{p['id']}</td><td>{p['title']}</td><td>{p['threat_actor']}</td><td>{len(json.loads(p['ttps_json'] or '[]'))}</td></tr>" for p in plans)}
        </table>
      </div>
    </div>
    """
    return html_page("Emulation Plans", body)

# --------- Exercises ----------
@app.route("/exercises", methods=["GET","POST"])
def exercises():
    db = get_db()
    if request.method == "POST":
        f = request.form
        db.execute("""INSERT INTO exercises (title,description,env,start_date,end_date,objectives,status,related_roe_id,related_plan_id,created_at)
                      VALUES (?,?,?,?,?,?,?,?,?,?)""",
                   (f.get("title"), f.get("description"), f.get("env"), f.get("start_date"), f.get("end_date"),
                    f.get("objectives"), f.get("status"), f.get("related_roe_id") or None,
                    f.get("related_plan_id") or None, now_iso()))
        db.commit()
        return redirect(url_for("exercises"))
    rows = rows_to_list(db.execute("SELECT * FROM exercises ORDER BY id DESC"))
    roes = rows_to_list(db.execute("SELECT id,title FROM roe ORDER BY id DESC"))
    plans = rows_to_list(db.execute("SELECT id,title FROM emulation_plans ORDER BY id DESC"))
    opt_roe = "".join([f"<option value='{r['id']}'>{r['id']} • {r['title']}</option>" for r in roes]) or "<option value=''>None</option>"
    opt_plan = "".join([f"<option value='{p['id']}'>{p['id']} • {p['title']}</option>" for p in plans]) or "<option value=''>None</option>"
    # Stakeholders quick view per exercise (count)
    counts = {r["id"]: r["c"] for r in db.execute("SELECT exercise_id id, COUNT(*) c FROM assignments GROUP BY exercise_id").fetchall()}
    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">New Exercise</div>
        <input name="title" placeholder="Exercise title" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <textarea name="description" placeholder="Description" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <input name="env" placeholder="Environment (Prod/Staging/Lab)" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <div class="grid-2 mb-2">
          <input name="start_date" type="date" class="p-2 rounded bg-transparent border"/>
          <input name="end_date" type="date" class="p-2 rounded bg-transparent border"/>
        </div>
        <textarea name="objectives" placeholder="Objectives" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <select name="status" class="w-full mb-2 p-2 rounded bg-transparent border">
          <option>Planned</option><option>Active</option><option>Closed</option>
        </select>
        <div class="grid-2 mb-2">
          <select name="related_roe_id" class="p-2 rounded bg-transparent border"><option value="">RoE (optional)</option>{opt_roe}</select>
          <select name="related_plan_id" class="p-2 rounded bg-transparent border"><option value="">Emulation Plan (optional)</option>{opt_plan}</select>
        </div>
        <div class="text-xs opacity-70">Assign people in the <b>People</b> page.</div>
        <button class="btn">Save Exercise</button>
      </form>
      <div class="card">
        <div class="text-lg font-semibold mb-2">Exercises</div>
        <table class="tbl w-full text-sm">
          <tr><th>ID</th><th>Title</th><th>Status</th><th>Period</th><th>Stakeholders</th><th>PDF</th></tr>
          {''.join(f"<tr><td>{r['id']}</td><td>{r['title']}</td><td>{r['status']}</td><td>{r['start_date']} → {r['end_date']}</td><td>{counts.get(r['id'],0)}</td><td><a class='btn' href='/report/{r['id']}'>Generate</a></td></tr>" for r in rows)}
        </table>
      </div>
    </div>
    """
    return html_page("Exercises", body)

# --------- Findings ----------
@app.route("/findings", methods=["GET","POST"])
def findings():
    db = get_db()
    if request.method == "POST":
        f = request.form
        db.execute("""INSERT INTO findings (exercise_id,title,severity,asset,description,evidence,remediation,owner,status,created_at)
                      VALUES (?,?,?,?,?,?,?,?,?,?)""",
                   (f.get("exercise_id") or None, f.get("title"), f.get("severity"), f.get("asset"),
                    f.get("description"), f.get("evidence"), f.get("remediation"),
                    f.get("owner"), f.get("status"), now_iso()))
        db.commit()
        return redirect(url_for("findings"))
    exs = rows_to_list(get_db().execute("SELECT id,title FROM exercises ORDER BY id DESC"))
    opt_ex = "<option value=''>Unlinked</option>" + "".join([f"<option value='{e['id']}'>#{e['id']} • {e['title']}</option>" for e in exs])
    rows = rows_to_list(db.execute("SELECT * FROM findings ORDER BY id DESC"))
    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">Submit Finding</div>
        <input name="title" placeholder="Finding title" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <div class="grid-2 mb-2">
          <select name="severity" class="p-2 rounded bg-transparent border">
            <option>Critical</option><option>High</option><option>Medium</option><option>Low</option><option>Info</option>
          </select>
          <select name="status" class="p-2 rounded bg-transparent border">
            <option>Open</option><option>In Progress</option><option>Closed</option>
          </select>
        </div>
        <input name="asset" placeholder="Impacted asset (host/app/account)" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <textarea name="description" placeholder="Description" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <textarea name="evidence" placeholder="Evidence / PoC" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <textarea name="remediation" placeholder="Remediation" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <div class="grid-2 mb-2">
          <input name="owner" placeholder="Owner" class="p-2 rounded bg-transparent border"/>
          <select name="exercise_id" class="p-2 rounded bg-transparent border">{opt_ex}</select>
        </div>
        <button class="btn">Save Finding</button>
      </form>
      <div class="card">
        <div class="text-lg font-semibold mb-2">Findings</div>
        <table class="tbl w-full text-sm">
          <tr><th>ID</th><th>Title</th><th>Sev</th><th>Status</th><th>Asset</th></tr>
          {''.join(f"<tr><td>{r['id']}</td><td>{r['title']}</td><td>{r['severity']}</td><td>{r['status']}</td><td>{r['asset']}</td></tr>" for r in rows)}
        </table>
      </div>
    </div>
    """
    return html_page("Findings", body)

# --------- IoCs ----------
@app.route("/iocs", methods=["GET","POST"])
def iocs():
    db = get_db()
    if request.method == "POST":
        f = request.form
        db.execute("""INSERT INTO iocs (type,value,related_technique,notes,created_at)
                      VALUES (?,?,?,?,?)""",
                   (f.get("type"), f.get("value"), f.get("related_technique"), f.get("notes"), now_iso()))
        db.commit()
        return redirect(url_for("iocs"))
    rows = rows_to_list(db.execute("SELECT * FROM iocs ORDER BY id DESC"))
    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">Add IoC</div>
        <div class="grid-2 mb-2">
          <select name="type" class="p-2 rounded bg-transparent border">
            <option>hash</option><option>ip</option><option>domain</option><option>url</option><option>email</option><option>path</option>
          </select>
          <input name="value" placeholder="Value" class="p-2 rounded bg-transparent border"/>
        </div>
        <input name="related_technique" placeholder="Related Technique (e.g., T1059)" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <textarea name="notes" placeholder="Notes" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <button class="btn">Save IoC</button>
      </form>
      <div class="card">
        <div class="text-lg font-semibold mb-2">IoC List</div>
        <table class="tbl w-full text-sm">
          <tr><th>ID</th><th>Type</th><th>Value</th><th>TTP</th></tr>
          {''.join(f"<tr><td>{r['id']}</td><td>{r['type']}</td><td>{r['value']}</td><td>{r['related_technique']}</td></tr>" for r in rows)}
        </table>
      </div>
    </div>
    """
    return html_page("Indicators of Compromise", body)

# --------- Threat Intel ----------
@app.route("/intel", methods=["GET","POST"])
def intel():
    db = get_db()
    if request.method == "POST":
        f = request.form
        db.execute("""INSERT INTO intel (title,source,tags,notes,created_at) VALUES (?,?,?,?,?)""",
                   (f.get("title"), f.get("source"), f.get("tags"), f.get("notes"), now_iso()))
        db.commit()
        return redirect(url_for("intel"))
    rows = rows_to_list(db.execute("SELECT * FROM intel ORDER BY id DESC"))
    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">Add Intel Note</div>
        <input name="title" placeholder="Title" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <div class="grid-2 mb-2">
          <input name="source" placeholder="Source (URL/Vendor)" class="p-2 rounded bg-transparent border"/>
          <input name="tags" placeholder="Tags (comma-separated)" class="p-2 rounded bg-transparent border"/>
        </div>
        <textarea name="notes" placeholder="Notes / Summary" class="w-full mb-2 p-2 rounded bg-transparent border h-24"></textarea>
        <button class="btn">Save</button>
      </form>
      <div class="card">
        <div class="text-lg font-semibold mb-2">Intel</div>
        <table class="tbl w-full text-sm">
          <tr><th>ID</th><th>Title</th><th>Source</th><th>Tags</th></tr>
          {''.join(f"<tr><td>{r['id']}</td><td>{r['title']}</td><td>{r['source']}</td><td>{r['tags']}</td></tr>" for r in rows)}
        </table>
      </div>
    </div>
    """
    return html_page("Threat Intelligence", body)

# --------- MITRE Map ----------
@app.route("/mitre")
def mitre():
    db = get_db()
    q = request.args.get("q","").strip().lower()

    rows = rows_to_list(db.execute(
        "SELECT technique_id,tactic,name,description FROM ttps ORDER BY tactic,name,technique_id"
    ))
    if q:
        def _match(r):
            return (
                q in (r["name"] or "").lower() or
                q in (r["technique_id"] or "").lower() or
                q in (r["tactic"] or "").lower()
            )
        rows = [r for r in rows if _match(r)]

    groups = defaultdict(list)
    for r in rows:
        groups[r["tactic"]].append(r)

    grid = ""
    for tac in TACTIC_ORDER:
        items = groups.get(tac, [])
        pills = "".join(
            f"<div class='pill' title='{esc_attr(r['description'] or '')}'>"
            f"{(r['name'] or r['technique_id'])} "
            f"<span class='opacity-70'>({r['technique_id']})</span></div>"
            for r in items
        ) or "<i class='opacity-60'>No techniques loaded</i>"
        grid += (
            f"<div class='card'>"
            f"<div class='font-semibold mb-1'>{TACTIC_TITLES[tac]} "
            f"<span class='opacity-60 text-xs'>({len(items)})</span></div>"
            f"<div class='flex flex-wrap gap-2'>{pills}</div>"
            f"</div>"
        )

    search_box = (
        "<form method='get' class='mb-3'>"
        f"<input class='search' name='q' placeholder='Search by name, ID or tactic (e.g., T1059, initial-access)' value='{esc_attr(q)}'/>"
        "</form>"
    )

    body = f"{search_box}<div class='grid-2'>{grid}</div>"
    return html_page("MITRE ATT&CK Map", body)

# --------- Kill Chain ----------
KILL_CHAIN_STAGES = [
    "Reconnaissance","Weaponization","Delivery","Exploitation","Installation",
    "Command & Control","Actions on Objectives"
]
@app.route("/killchain", methods=["GET","POST"])
def killchain():
    db = get_db()
    all_ttps = rows_to_list(db.execute("SELECT technique_id,name,tactic FROM ttps ORDER BY tactic,name"))
    options = "".join([
        f"<option value='{t['technique_id']}|{t['name'] or t['technique_id']}|{t['tactic']}'>"
        f"{TACTIC_TITLES.get(t['tactic'], t['tactic'].title())} • {(t['name'] or t['technique_id'])} ({t['technique_id']})"
        f"</option>"
        for t in all_ttps
    ])

    if request.method == "POST":
        f = request.form
        mapping = {}
        for stage in KILL_CHAIN_STAGES:
            mapping[stage] = [json.loads(x) for x in json.loads(f.get(stage, "[]"))]
        db.execute("INSERT INTO killchain (name,mapping_json,created_at) VALUES (?,?,?)",
                   (f.get("name"), json.dumps(mapping), now_iso()))
        db.commit()
        return redirect(url_for("killchain"))

    chains = rows_to_list(db.execute("SELECT * FROM killchain ORDER BY id DESC"))
    stage_blocks = ""
    for stage in KILL_CHAIN_STAGES:
        stage_blocks += f"""
        <div class="card">
          <div class="font-semibold mb-1">{stage}</div>
          <select id="sel_{stage}" multiple size="6" class="w-full p-2 rounded bg-transparent border">{options}</select>
          <input type="hidden" id="hid_{stage}" name="{stage}"/>
          <button type="button" class="btn mt-2" onclick="
            const sel=[...document.getElementById('sel_{stage}').selectedOptions].map(o=>JSON.parse(o.value));
            document.getElementById('hid_{stage}').value = JSON.stringify(sel);
            alert('{stage}: staged '+sel.length+' TTP(s).');
          ">Stage</button>
        </div>"""

    chains_table = "".join([
        f"<tr><td>{c['id']}</td><td>{c['name']}</td><td><button class='btn' hx-get='/killchain/view/{c['id']}' hx-target='#kcview'>View</button></td></tr>"
        for c in chains
    ])

    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">New Kill Chain</div>
        <input name="name" placeholder="Name" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <div class="grid-2">{stage_blocks}</div>
        <button class="btn mt-3">Save Kill Chain</button>
      </form>
      <div class="">
        <div class="card mb-3">
          <div class="text-lg font-semibold mb-2">Saved Chains</div>
          <table class="tbl w-full text-sm"><tr><th>ID</th><th>Name</th><th>View</th></tr>{chains_table}</table>
        </div>
        <div id="kcview" class="card"><i>Select a chain to view…</i></div>
      </div>
    </div>
    """
    return html_page("Cyber Kill Chain Builder", body)

@app.get("/killchain/view/<int:kc_id>")
def view_kc(kc_id):
    db = get_db()
    c = db.execute("SELECT * FROM killchain WHERE id=?", (kc_id,)).fetchone()
    if not c:
        return "<div class='text-red-500'>Not found.</div>"
    mapping = json.loads(c["mapping_json"] or "{}")
    blocks = ""
    for stage in KILL_CHAIN_STAGES:
        items = mapping.get(stage, [])
        pills = ""
        for x in items:
            if isinstance(x, (list, tuple)) and len(x) >= 2:
                tid, name = x[0], x[1]
            elif isinstance(x, dict):
                tid = x.get("0") or x.get("technique_id") or "TTP"
                name = x.get("1") or x.get("name") or "Name"
            else:
                tid, name = "TTP", "Name"
            pills += f"<div class='chip bg-slate-500/20'>{name} ({tid})</div>"
        blocks += f"<div class='card'><div class='font-semibold'>{stage}</div><div class='flex flex-wrap gap-2 mt-2'>{pills or '<i>None</i>'}</div></div>"
    html = f"<div class='text-lg font-semibold mb-2'>Kill Chain: {c['name']}</div><div class='grid-2'>{blocks}</div>"
    resp = make_response(html)
    resp.headers["HX-Reswap"] = "outerHTML"
    return resp

# --------- People & Assignments ----------
@app.route("/people", methods=["GET","POST"])
def people():
    db = get_db()
    if request.method == "POST":
        f = request.form
        if f.get("form") == "person":
            db.execute("""INSERT INTO people (name,role,contact,notes,created_at)
                          VALUES (?,?,?,?,?)""",
                       (f.get("name"), f.get("role"), f.get("contact"), f.get("notes"), now_iso()))
            db.commit()
        elif f.get("form") == "assign":
            db.execute("""INSERT INTO assignments (exercise_id,person_id,duty,created_at)
                          VALUES (?,?,?,?)""",
                       (f.get("exercise_id"), f.get("person_id"), f.get("duty"), now_iso()))
            db.commit()
        return redirect(url_for("people"))

    persons = rows_to_list(db.execute("SELECT * FROM people ORDER BY id DESC"))
    exercises_list = rows_to_list(db.execute("SELECT id,title FROM exercises ORDER BY id DESC"))
    # Show current assignments with names
    assigs = rows_to_list(db.execute("""
        SELECT a.id, a.exercise_id, a.person_id, a.duty,
               e.title AS exercise, p.name AS person, p.role AS role
        FROM assignments a
        LEFT JOIN exercises e ON e.id=a.exercise_id
        LEFT JOIN people p ON p.id=a.person_id
        ORDER BY a.id DESC
    """))
    opt_people = "".join([f"<option value='{p['id']}'>{p['name']} • {p['role']}</option>" for p in persons]) or "<option disabled>No people</option>"
    opt_ex    = "".join([f"<option value='{e['id']}'>#{e['id']} • {e['title']}</option>" for e in exercises_list]) or "<option disabled>No exercises</option>"

    body = f"""
    <div class="grid-2">
      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">Add Person</div>
        <input type="hidden" name="form" value="person"/>
        <input name="name" placeholder="Full name" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <input name="role" placeholder="Role (e.g., RT Lead, Blue POC)" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <input name="contact" placeholder="Contact (email/phone/Slack)" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <textarea name="notes" placeholder="Notes" class="w-full mb-2 p-2 rounded bg-transparent border h-20"></textarea>
        <button class="btn">Save Person</button>
      </form>

      <form method="post" class="card">
        <div class="text-lg font-semibold mb-2">Assign to Exercise</div>
        <input type="hidden" name="form" value="assign"/>
        <div class="grid-2 mb-2">
          <select name="person_id" class="p-2 rounded bg-transparent border">{opt_people}</select>
          <select name="exercise_id" class="p-2 rounded bg-transparent border">{opt_ex}</select>
        </div>
        <input name="duty" placeholder="Duty / responsibility (e.g., Approval, Comms, Intel)" class="w-full mb-2 p-2 rounded bg-transparent border"/>
        <button class="btn">Add Assignment</button>
      </form>
    </div>

    <div class="card mt-4">
      <div class="text-lg font-semibold mb-2">Assignments</div>
      <table class="tbl w-full text-sm">
        <tr><th>ID</th><th>Exercise</th><th>Person</th><th>Role</th><th>Duty</th></tr>
        {''.join(f"<tr><td>{a['id']}</td><td>#{a['exercise_id']} • {a.get('exercise') or '-'}</td><td>{a.get('person') or '-'}</td><td>{a.get('role') or '-'}</td><td>{a.get('duty') or '-'}</td></tr>" for a in assigs)}
      </table>
    </div>
    """
    return html_page("People & Assignments", body)

# --------- Reports (PDF) ----------
@app.get("/reports")
def reports():
    db = get_db()
    exs = rows_to_list(db.execute("SELECT id,title,status,start_date,end_date FROM exercises ORDER BY id DESC"))
    trows = "".join([f"<tr><td>{e['id']}</td><td>{e['title']}</td><td>{e['status']}</td><td>{e['start_date']} → {e['end_date']}</td><td><a class='btn' href='/report/{e['id']}'>Generate PDF</a></td></tr>" for e in exs])
    body = f"""
    <div class="card">
      <div class="text-lg font-semibold mb-2">Generate Exercise Report (PDF)</div>
      <table class="tbl w-full text-sm">
        <tr><th>ID</th><th>Title</th><th>Status</th><th>Period</th><th>PDF</th></tr>
        {trows}
      </table>
    </div>
    """
    return html_page("Reports", body)

@app.get("/report/<int:exercise_id>")
def report(exercise_id):
    db = get_db()
    e = db.execute("SELECT * FROM exercises WHERE id=?", (exercise_id,)).fetchone()
    if not e:
        return html_page("Report", "<div class='card text-red-500'>Exercise not found.</div>")
    roe = db.execute("SELECT * FROM roe WHERE id=?", (e["related_roe_id"],)).fetchone() if e["related_roe_id"] else None
    plan = db.execute("SELECT * FROM emulation_plans WHERE id=?", (e["related_plan_id"],)).fetchone() if e["related_plan_id"] else None
    fnds = rows_to_list(db.execute("SELECT * FROM findings WHERE exercise_id=? ORDER BY severity DESC", (exercise_id,)))
    iocs = rows_to_list(db.execute("SELECT * FROM iocs ORDER BY id DESC"))
    stakeholders = rows_to_list(db.execute("""
        SELECT p.name, p.role, a.duty, p.contact
        FROM assignments a LEFT JOIN people p ON p.id=a.person_id
        WHERE a.exercise_id=? ORDER BY p.name
    """, (exercise_id,)))

    buff = BytesIO()
    c = rl_canvas.Canvas(buff, pagesize=A4)
    W, H = A4

    def H1(txt, y): c.setFont("Helvetica-Bold", 16); c.drawString(2*cm, y, txt); return y-0.8*cm
    def H2(txt, y): c.setFont("Helvetica-Bold", 12); c.drawString(2*cm, y, txt); return y-0.6*cm
    def P(txt, y):
        c.setFont("Helvetica", 10)
        for line in (txt or "-").splitlines():
            c.drawString(2*cm, y, line); y -= 0.45*cm
        return y-0.2*cm
    def KV(key, val, y): c.setFont("Helvetica-Bold", 10); c.drawString(2*cm, y, f"{key}:"); c.setFont("Helvetica", 10); c.drawString(4.5*cm, y, (val or "-")[:85]); return y-0.45*cm

    y = H - 2*cm
    y = H1(f"{APP_NAME} • Exercise Report #{e['id']}", y)
    y = KV("Title", e["title"], y)
    y = KV("Status", e["status"], y)
    y = KV("Period", f"{e['start_date']} → {e['end_date']}", y)
    y = KV("Environment", e["env"], y)
    y = H2("Objectives", y); y = P(e["objectives"])
    y = H2("Description", y); y = P(e["description"])

    if stakeholders:
        if y < 4*cm: c.showPage(); y = H-2*cm
        y = H1("Stakeholders", y)
        for s in stakeholders:
            y = P(f"- {s['name']} • {s['role']} • {s['duty']} • {s['contact']}", y)
            if y < 3*cm: c.showPage(); y = H-2*cm

    if roe:
        if y < 4*cm: c.showPage(); y = H-2*cm
        y = H1("Rules of Engagement (linked)", y)
        y = KV("Title", roe["title"], y)
        y = KV("Customer", roe["customer"], y)
        y = H2("Rules", y); y = P(roe["rules"])

    if plan:
        if y < 4*cm: c.showPage(); y = H-2*cm
        y = H1("Emulation Plan (linked)", y)
        y = KV("Title", plan["title"], y)
        y = KV("Threat Actor", plan["threat_actor"], y)
        try:
            ttps = json.loads(plan["ttps_json"] or "[]")
        except:
            ttps = []
        y = H2("Selected TTPs", y)
        if not ttps:
            y = P("-", y)
        else:
            for t in ttps:
                if isinstance(t, list):
                    tid, tac, nm = (t[0] if len(t)>0 else "?"), (t[1] if len(t)>1 else "?"), (t[2] if len(t)>2 else "?")
                elif isinstance(t, dict):
                    tid = t.get("0") or t.get("technique_id") or "?"
                    tac = t.get("1") or t.get("tactic") or "?"
                    nm  = t.get("2") or t.get("name") or "?"
                else:
                    tid, tac, nm = "?", "?", "?"
                y = P(f"- {tid} • {nm} [{TACTIC_TITLES.get(tac, tac)}]", y)
                if y < 3*cm: c.showPage(); y = H-2*cm

    if y < 4*cm: c.showPage(); y = H-2*cm
    y = H1("Findings", y)
    if not fnds:
        y = P("No findings submitted.", y)
    else:
        for f in fnds:
            y = H2(f"#{f['id']} • {f['title']}  ({f['severity']})", y)
            y = KV("Asset", f["asset"], y)
            y = KV("Status", f["status"], y)
            y = KV("Owner", f["owner"], y)
            y = H2("Description", y); y = P(f["description"])
            y = H2("Evidence", y); y = P(f["evidence"])
            y = H2("Remediation", y); y = P(f["remediation"])
            if y < 3*cm: c.showPage(); y = H-2*cm

    if y < 4*cm: c.showPage(); y = H-2*cm
    y = H1("IoCs (Snapshot)", y)
    show = iocs[:30]
    for i in show:
        y = P(f"- [{i['type']}] {i['value']}  (TTP: {i['related_technique'] or '-'})", y)
        if y < 3*cm: c.showPage(); y = H-2*cm

    c.showPage(); c.save(); buff.seek(0)
    filename = f"Exercise_{exercise_id}_Report.pdf"
    return send_file(buff, as_attachment=True, download_name=filename, mimetype="application/pdf")

# --------- Admin (ATT&CK Import/Enrich) ----------
@app.get("/admin")
def admin():
    mitre_loaded = "Yes" if get_setting("mitre_loaded", "0") == "1" else "No"
    body = f"""
    <div class="grid-2">
      <div class="card">
        <div class="text-lg font-semibold mb-2">MITRE ATT&CK Import</div>
        <div class="text-sm mb-2">Status: <b>{mitre_loaded}</b></div>
        <p class="text-sm opacity-80 mb-2">Reads <code>enterprise-attack.json</code> (STIX or Navigator layer) and/or <code>enterprise-attack-v17.1.xlsx</code>.</p>
        <a class="btn" href="/admin/load_mitre">Import / Enrich MITRE ATT&CK</a>
      </div>
      <div class="card">
        <div class="text-lg font-semibold mb-2">Tips</div>
        <ul class="list-disc pl-5 text-sm">
          <li>Use <b>People</b> to add stakeholders and assign to exercises.</li>
          <li>Emulation Plans: pick TTPs by tactic with names from ATT&CK.</li>
          <li>Dashboard charts auto-update from Findings.</li>
        </ul>
      </div>
    </div>
    """
    return html_page("Admin", body)

def _load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _try_import_from_stix_or_layer(db, js):
    """
    Supports:
      - STIX bundle: objects[].attack-pattern + kill_chain_phases (phase_name = tactic)
      - Navigator layer: techniques[{techniqueID, tactic}] (fills placeholders; enriched later via XLSX)
    """
    now = now_iso()
    inserted = 0
    # STIX bundle
    if isinstance(js, dict) and "objects" in js:
        for obj in js.get("objects", []):
            if obj.get("type") == "attack-pattern":
                name = obj.get("name")
                desc = (obj.get("description") or "").strip()
                tech_id = None
                for ref in obj.get("external_references", []):
                    if ref.get("source_name","").lower() == "mitre-attack" and str(ref.get("external_id","")).startswith("T"):
                        tech_id = ref["external_id"]; break
                if not tech_id:
                    continue
                tactics = [p.get("phase_name") for p in obj.get("kill_chain_phases", []) if p.get("kill_chain_name") == "mitre-attack"]
                if not tactics:
                    continue
                for tac in tactics:
                    tac = (tac or "").lower()
                    db.execute("""INSERT OR IGNORE INTO ttps (technique_id,tactic,name,description,refs,created_at)
                                  VALUES (?,?,?,?,?,?)""",
                               (tech_id, tac, name, desc, "", now))
                    inserted += 1
        db.commit()
        return inserted

    # Navigator layer (IDs + tactic only)
    if isinstance(js, dict) and "techniques" in js:
        for t in js["techniques"]:
            tech_id = t.get("techniqueID")
            tac = (t.get("tactic") or "").lower()
            if not tech_id or not tac:
                continue
            db.execute("""INSERT OR IGNORE INTO ttps (technique_id,tactic,name,description,refs,created_at)
                          VALUES (?,?,?,?,?,?)""",
                       (tech_id, tac, tech_id, "", "", now))
            inserted += 1
        db.commit()
        return inserted
    return 0

def _try_enrich_from_xlsx(db, xlsx_path):
    """Enrich names/desc per tactic from ATT&CK Excel (robust header/sheet scan)."""
    try:
        import openpyxl
    except Exception:
        return 0
    if not os.path.exists(xlsx_path):
        return 0

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    updated = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Build header map (fuzzy)
        header = {}
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=1, column=c).value
            if v is None: continue
            k = str(v).strip().lower()
            header[k] = c

        # Find candidate columns by fuzzy contains
        def find_col(*keys):
            for k,c in header.items():
                for key in keys:
                    if key in k:
                        return c
            return None

        col_id   = find_col("technique id", "external id", "id", "external_id")
        col_name = find_col("technique name", "technique", "name")
        col_desc = find_col("description", "technique description")
        col_tac  = find_col("tactics", "tactic", "domain tactics")

        if not col_id or not col_name or not col_tac:
            continue  # not a techniques-like sheet

        for r in range(2, ws.max_row+1):
            tid = str(ws.cell(row=r, column=col_id).value or "").strip()
            if not tid.startswith("T"): 
                continue
            nm = (ws.cell(row=r, column=col_name).value or "").strip()
            ds = (ws.cell(row=r, column=col_desc).value or "").strip() if col_desc else ""
            tacs_raw = (ws.cell(row=r, column=col_tac).value or "").strip()
            if not tacs_raw: 
                continue
            # Split tactics robustly
            tacs = str(tacs_raw).replace("/", ",").replace(" & ", ",").split(",")
            for tac in [t.strip().lower().replace(" ", "-") for t in tacs]:
                if tac not in TACTIC_TITLES: 
                    continue
                # insert or update
                db.execute("""INSERT OR IGNORE INTO ttps (technique_id,tactic,name,description,refs,created_at)
                              VALUES (?,?,?,?,?,?)""",
                           (tid, tac, nm, ds, "", now_iso()))
                db.execute("""UPDATE ttps 
                              SET name = CASE WHEN name IS NULL OR name='' OR name = technique_id THEN ? ELSE name END,
                                  description = CASE WHEN description IS NULL OR description='' THEN ? ELSE description END
                              WHERE technique_id=? AND tactic=?""",
                           (nm, ds, tid, tac))
                updated += 1
    db.commit()
    return updated

@app.get("/admin/load_mitre")
def admin_load_mitre():
    db = get_db()
    total = 0
    # Import base (JSON)
    for p in MITRE_JSON_CANDIDATES:
        if os.path.exists(p):
            try:
                js = _load_json(p)
                total += _try_import_from_stix_or_layer(db, js)
            except Exception:
                pass
    # Enrich with XLSX
    for p in MITRE_XLSX_CANDIDATES:
        if os.path.exists(p):
            try:
                total += _try_enrich_from_xlsx(db, p)
            except Exception:
                pass

    if total > 0:
        set_setting("mitre_loaded", "1")
        msg = f"Imported/updated {total} tactic-technique rows."
    else:
        msg = "No ATT&CK data imported (files missing or unrecognized)."

    body = f"""
    <div class="card">
      <div class="text-lg font-semibold mb-2">MITRE ATT&CK Import</div>
      <p class="mb-2">{msg}</p>
      <a class="btn" href="/mitre">Go to MITRE Map</a>
      <a class="btn" href="/admin">Back</a>
    </div>
    """
    return html_page("Admin • Import", body)

# --------------------- Main ---------------------
if __name__ == "__main__":
    app.run(debug=True)
